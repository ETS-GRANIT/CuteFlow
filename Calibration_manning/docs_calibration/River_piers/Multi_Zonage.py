'''Script de zonage
==========================================================================
Lit le fichier exel contenant les cellules et nom des zones sélectionnées
puis génères des dossiers Case_%d contenant chacun le fichier
zonage_case_%d.xlsx'
..........................................................................
INPUTS :
  * Manning.txt :  fichier de sortie de LHS.py contenant les nombres de
                   Manning pour les N-uplets échantillons
  * zone.xlsx :    fichier excel contenant les noms des "m" zones (zone1,
                   zone2,...,zonem) et les cellules correspondantes
 OUTPUTS :
  * zonage_case_%d.xlsx :         un ensemble de N fichiers contenant en
                                  plus des zones et cellules les nombres
                                  de manning correspondants.
 * Case_%d. :                     un ensemble de N dossiers contenant
                                  les fichiers de zonage
========================================================================='''

#Importation des bibliothèques
import numpy as np
import os
import sys
import pandas as pd
import openpyxl

# Vérification du nom du fichier Excel de zonage
if len(sys.argv) < 2:
    print("Usage: python3 Zonage.py nom_du_fichier.xlsx")
    sys.exit(1)

# Chemin d'accès au fichier Excel de zonage
zone_filename = sys.argv[1]

# Lecture du fichier Excel
df = pd.read_excel(zone_filename,engine='openpyxl')

headers = df.columns.to_list()
zones = df['Zone'].to_list()
cells_IDS = df['Cells IDs'].to_list()

# Création de la bibliothèque des zones
zones_set=[]

for name_zone in zones:
    if name_zone not in zones_set:
        zones_set.append(name_zone)

# Chemnin d'accès fichiers des valeurs de Manning (Output du fichier LHS)
manning_filename = 'Manning.txt'
delimiter = ','

# Extraction de l'ensemble des valeurs de Manning
manning_set = np.genfromtxt(manning_filename, delimiter=delimiter)

# Nombre de fichier de sortie à générer
num_file = manning_set.shape[0]

# Formatage de l'entête de chaque fichier
headers.append('Manning')

# Création des nouveaux fichiers et écriture des nombres de manning dans chaque fichier
for id_file_manning in range(1, num_file + 1):  # Identifiant du nouveau fichier
    id_folder = id_file_manning                 # Identifiant du nouveau dossier
    Manning_val = []

    # Boucle d'affectation des valeurs de Manning
    for k in range(len(zones)):
        test_zone=zones[k]
        if test_zone in zones_set:
            index = zones_set.index(test_zone)
            Manning_val.append(manning_set[id_file_manning - 1][index])

    # Boucle de création des nouveaux dossiers et nouveaux fichiers

    # Générer le nom du nouveau dossier
    nom_dossier = f'Case_{id_folder}'

    # Vérification de l'existence du dossier
    if not os.path.exists(nom_dossier):
        # Création du dossier dans le répertoire actuel du script
        os.makedirs(nom_dossier)
    else:
        print(f'\nDossier {nom_dossier} existe déjà.')

    # Générer le chemin d'accès du fichier excel
    nom_fichier_excel = os.path.join(nom_dossier, f'zonage_{nom_dossier}.xlsx')

    # Création d'un nouveau fichier Excel
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Écriture de l'en-tête
    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col, value=header)

    # Formatage du nouveau fichier : Écriture des zones, cells_IDs et Manning_val
    for row, (zone, num, manning) in enumerate(zip(zones, cells_IDS, Manning_val), 2):
        sheet.cell(row=row, column=1, value=zone)
        sheet.cell(row=row, column=2, value=num)
        sheet.cell(row=row, column=3, value=manning)

    # Sauvegarde du fichier Excel
    wb.save(nom_fichier_excel)

    # Affichage du message d'évolution de la tâche
    print(f'Données modifiées enregistrées avec succès dans le fichier : {nom_fichier_excel}\n')

# Affichage du message de fin d'exécution
print('\nFin du programme!!\n')
