# Importation des bibliothèques
import openpyxl
import sys
import os

if(len(sys.argv) <= 1) :
    print("[-] Erreur: Pas assez d'arguments d'entrée")
    print("[-] Utilisation: python3 gen_Entree_gen_manning.py input_file.xlsx")
    exit(0)

# Chemin vers le fichier Excel
filename = sys.argv[1]

# Ouverture du fichier Excel 
workbook = openpyxl.load_workbook(filename)
sheet = workbook.active

# Extraction des données de la colonne 0 (cellIDs) et 1 (manningValues)
cellIDs = [int(sheet.cell(row=i, column=1).value) for i in range(2, sheet.max_row + 1)]

# Nettoyage et conversion des valeurs de la colonne 1 (manningValues)
manningValues = []
for i in range(2, sheet.max_row + 1):
    value = sheet.cell(row=i, column=2).value
    if isinstance(value, str):
        value = value.strip()
    manningValues.append(float(value))

# Écriture des données dans un fichier CSV
outputFilename = 'Entree_gen_manning.txt'
with open(outputFilename, 'w') as f:
    f.write('Cells IDs\n')
    f.write(','.join(str(id) for id in cellIDs) + '\n')

    f.write('Manning_values\n')
    f.write(','.join(str(value) for value in manningValues) + '\n')

print(f'Fichier CSV généré avec succès, avec: {len(cellIDs)} (IDs des cellules) et {len(manningValues)} (Valeurs de Manning)')