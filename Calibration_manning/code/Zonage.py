'''Script de zonage
==========================================================================
Lit le fichier exel contenant les cellules et nom des zones sélectionnées
puis génères le fichier Excel de zonage contenant les valeurs des coeff
de Manning.
..........................................................................
INPUTS :
  * Manning.txt :  fichier contenant les nombres de Manning pour les 'm'
                    zones identifiée
  * zone.xlsx :    fichier excel contenant les noms des "m" zones (zone1,
                   zone2,...,zonem) et les cellules correspondantes
 OUTPUT :
  * Output_zonage.xlsx
========================================================================='''

# Importation des bibliothèques
import numpy as np
import os
import sys
import pandas as pd
import openpyxl

# Vérification de la syntaxe
if len(sys.argv) < 2 :
    print("Usage : python3 Zonage.py nom_du_fichier.xlsx")
    sys.exit(1)

# Chemin d'accès au fichier Excel de zonage
zone_filename = sys.argv[1]

# Lecture du fichier Excel
df = pd.read_excel(zone_filename,engine='openpyxl')

headers = df.columns.to_list()
zones = df['Zone'].to_list()
cells_IDS = df['Cells IDs'].to_list()

# Création de la bibliothèque des zones
zones_set=[]
for name_zone in zones:
    if name_zone not in zones_set:
        zones_set.append(name_zone)

# Chemin d'accès au fichier de valeurs de Manning
manning_filename = 'Manning.txt'
delimiter = ','

# Extraction des valeurs de Manning
manning_set=np.genfromtxt(manning_filename,delimiter=delimiter)

# Actualisation des variables d'en-tête
headers.append('Manning')

# Liste des valeurs de Manning
Manning_val = []

# Boucle d'affectation des valeurs de Manning
for k in range(len(zones)):
    test_zone=zones[k]
    if test_zone in zones_set:
        index = zones_set.index(test_zone)
        Manning_val.append(manning_set[index])

# Création et écriture du fichier de sortie contenant les nombres de Manning
classeur = openpyxl.Workbook()
sheet = classeur.active

# Formatage du fichier de sortie : écriture des zones, cells_IDs et Manning_val
for col, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col, value=header)

for row, (zone, num, manning) in enumerate(zip(zones,cells_IDS,Manning_val),2):
    sheet.cell(row=row, column=1, value=zone)
    sheet.cell(row=row, column=2, value=num)
    sheet.cell(row=row, column=3, value=manning)

classeur.save('Output_zonage.xlsx')
print('\nFichier Excel créé et enregistré avec succès.\n')
